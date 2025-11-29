import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import util.conf.config as config
import util.util as util


class ToExcel:
    _instance = None

    def __new__(cls, *args, **kwargs):
        if not cls._instance:
            cls._instance = super().__new__(cls)
            cls._instance._initialized = False
        return cls._instance

    def __init__(self):
        if self._initialized:
            return
        self.excel_dir = config.output_dir()
        os.makedirs(self.excel_dir, exist_ok=True)
        self.file_path = config.excel_conf()['file_path']
        self.is_flush = config.excel_conf()['is_flush']
        self.is_write_excel = config.excel_conf()['is_write_excel']
        self.data = []
        self.reset()
        self._initialized = True

    def init(self):
        if self.is_write_excel:
            return
        if os.path.exists(self.file_path):
            util.rename_file(self.file_path)

    def reset(self):
        self.stock_name = None
        self.stock_code = None
        self.portfolio = None
        self.final_portfolio = None
        self.position_status = None
        self.gross = None
        self.yield_rate = None
        self.hold_size = None
        self.cost_price = None
        self.current_price = None
        self.cumulative_return = None
        self.max_drawdown = None
        self.sharpe_ratio = None

    def _create_record(self, **kwargs):
        """Helper method to create a standardized record dictionary"""
        record = {
            '股票名称': self.stock_name,
            '代码': self.stock_code,
            '启动资金': f"{self.portfolio:.2f}" if self.portfolio is not None else None,
            '时间': kwargs.get('time'),
            '操作': kwargs.get('op_type'),
            '价格': kwargs.get('price'),
            '数量': abs(kwargs.get('size')) if kwargs.get('size') is not None else None,
            '交易金额': kwargs.get('value'),
            '交易佣金': f"{kwargs.get('comm'):.2f}" if kwargs.get('comm') is not None else None,
            '当前持仓': kwargs.get('hold_size'),
            # '累计成本': kwargs.get('acc_cost'),
            '成本价': f"{self.cost_price:.3f}" if self.cost_price is not None else None,
            '现价': f"{self.current_price}" if self.current_price is not None else None,
            '涨幅': f"{self.yield_rate:.2%}" if self.yield_rate is not None else None,
            '持仓状态': f"{self.position_status}\n({self.hold_size}股)" if self.position_status == "持仓中" else self.position_status,
            '最终资金': f"{self.final_portfolio:.2f}" if self.final_portfolio is not None else None,
            '总收益': f"{self.gross:.2f}" if self.gross is not None else None,
            '累计收益率': f"{self.cumulative_return:.2f}%" if self.cumulative_return is not None else None,
            '最大回撤': f"{self.max_drawdown:.2f}%" if self.max_drawdown is not None else None,
            '夏普比率': f"{self.sharpe_ratio:.2f}" if self.sharpe_ratio is not None else None,
        }
        return record

    def write_state(self, params):
        '''
        记录起始/终止态数据（需要合并单元格的字段的数据），但未写入表格。
        待每次股票处理结束后，调用write_finish写入
        '''
        for key, value in params.items():
            setattr(self, key, value)

    def write_finish(self):
        '''每只股票处理结束，将起始/终止态数据写入表格'''
        if self.stock_code is not None:
            state_record = self._create_record()
            self.data.append(state_record)
        self.reset()

    def write_signal(self, operation_type, price, time):
        '''记录买卖信号'''
        self.data.append(self._create_record(op_type=operation_type, price=price, time=time))

    def write_op_create(self, operation_type, price, size, time):
        '''记录买卖创建'''
        self.data.append(self._create_record(op_type=operation_type, price=price, size=size, time=time))

    def write_operation(self, operation_type, operation_detail, accumulated_cost, holding_size, time):
        '''记录买卖创建/执行'''

        if holding_size is not None:
            self.hold_size = holding_size

        op_price = operation_detail.price
        if operation_type == "SELL EXECUTED":
            op_price = operation_detail.pprice

        self.data.append(self._create_record(time=time,
                                             op_type=operation_type,
                                             price=op_price,
                                             size=operation_detail.size,
                                             value=operation_detail.value,
                                             comm=operation_detail.comm,
                                             hold_size=holding_size,
                                             acc_cost=accumulated_cost))

    def export_to_excel(self):
        '''每个stock_list_file处理结束后，写入xlsx文件'''
        if self.is_write_excel:
            return

        if not self.file_path:
            print(f"Error: export_to_excel: file_path is None")
            return

        if not self.data:
            print(f"Error: export_to_excel: data is empty")
            return

        df = pd.DataFrame(self.data)
        df.to_excel(self.file_path, index=False, engine='openpyxl')

        # 使用 openpyxl 合并单元格
        wb = load_workbook(self.file_path)
        ws = wb.active

        # 冻结首行
        ws.freeze_panes = "A2"

        # 需要合并的单元格的列编号
        merge_config = ['A', 'B', 'C', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']

        # 需要换行展示的单元格列编号
        wrap_config = ['N']

        # 获取所有股票代码
        stock_codes = df['代码'].unique()

        for code in stock_codes:
            # 找到该股票的所有行
            rows = [i for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True)) if row[1] == code]

            if len(rows) > 1:
                start_row = rows[0] + 1  # Excel 行号从 1 开始
                end_row = rows[-1] + 1

                # 获取每列最后一个非空值
                last_values = {}
                for col in merge_config:
                    for row_idx in reversed(range(start_row, end_row + 1)):
                        cell_value = ws[f"{col}{row_idx}"].value
                        if cell_value is not None and cell_value != "":
                            last_values[col] = cell_value
                            break

                # 合并单元格并设置值
                for col in merge_config:
                    ws.merge_cells(f"{col}{start_row}:{col}{end_row}")
                    if col in last_values:
                        ws[f"{col}{start_row}"].value = last_values[col]
                    ws[f"{col}{start_row}"].alignment = Alignment(
                        horizontal='center',
                        vertical='top'
                    )

                ws.merge_cells(f'A{start_row}:A{end_row}')
                ws.merge_cells(f'B{start_row}:B{end_row}')

                # 单元格内换行展示
                for col in wrap_config:
                    for row in range(2, ws.max_row + 1):
                        ws[f"{col}{row}"].alignment = Alignment(
                            wrap_text=True,  # 启用自动换行
                            vertical='top'  # 文字顶部对齐
                        )

        wb.save(self.file_path)


exceler = ToExcel()
